#!/usr/bin/python3.6
# -*- coding: UTF-8 -*-

import os, sys, io
import time
import openpyxl
from openpyxl.cell import cell

sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='utf-8')

_EXCELPATH_ = "/data/excel/ZentaoMaker/ZentaoMaker.xlsx"

def GetExcelInfo():
	wb = openpyxl.load_workbook(_EXCELPATH_)
	sheet = wb["Sheet1"]

	listTitle = []
	listMsg = []
	dataTitle = ""
	dataMsg = ""

	max_row = sheet.max_row
	max_column = sheet.max_column
	nIndex = 2
	if max_column > 7:
		nIndex = max_column - 7
	while nIndex < max_column:
		title = sheet.cell(row = 1, column = nIndex + 1).value
		msg = sheet.cell(row = 2, column = nIndex + 1).value
		dataTitle = dataTitle + "\"" + str(title) + "\","
		dataMsg = dataMsg + str(msg) + ","
		listTitle.append(title)
		listMsg.append(msg)
		nIndex = nIndex + 1
	if dataTitle != "":
		dataTitle = dataTitle[0: len(dataTitle) - 1]
	if dataMsg != "":
		dataMsg = dataMsg[0: len(dataMsg) - 1]
	#print(dataTitle)
	#print(dataMsg)
	return dataTitle, dataMsg

def PrintAll(dataTitle, dataMsg):
	print ("Content-type:text/html")
	print ("")
	print ("<!DOCTYPE html>")
	print ("<html>")
	print ("<head>")
	print ("<title>禅道BUG统计</title>")
	print ("	<meta charset=\"utf-8\">")
	print ("	<!-- 引入 ECharts 文件 -->")
	print ("	<script src=\"../echarts.min.js\"></script>")
	print ("</head>")
	print ("<body>")
	print ("")
	print ("<!-- 为 ECharts 准备一个具备大小（宽高）的 DOM -->")
	print ("	<div id=\"main\" style=\"height:400px;\"></div>")
	print ("	<script type=\"text/javascript\">")
	print ("		// 基于准备好的dom，初始化echarts实例")
	print ("		var myChart = echarts.init(document.getElementById('main'));")
	print ("")
	print ("		// 指定图表的配置项和数据")
	print ("		var option = {")
	print ("				title: {")
	print ("				text: '每日BUG数'")
	print ("			},")
	print ("			tooltip: {},")
	print ("			legend: {")
	print ("				data:['num']")
	print ("			},")
	print ("			xAxis: {")
	print ("				data: [" + dataTitle + "]")
	print ("			},")
	print ("			yAxis: {},")
	print ("			series: [{")
	print ("				name: '',")
	print ("				type: 'line',")
	print ("				data: [" + dataMsg + "]")
	print ("			}]")
	print ("		};")
	print ("")
	print ("		// 使用刚指定的配置项和数据显示图表。")
	print ("		myChart.setOption(option);")
	print ("	</script>")
	print ("")
	print ("<a href=\"index\">返回主页</a>")
	print ("</body>")
	print ("</html>")
	return

def main():
	dataTitle, dataMsg = GetExcelInfo()
	PrintAll(dataTitle, dataMsg)
	return

main()
